import requests
import PyPDF2
import docx
import openpyxl
import pandas as pd
from handle_error import log_error
from handle_visted_unvisted import save_link


def extract_links_from_pdf(file_path):
    links = set()
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                if '/Annots' in page:
                    annots = page['/Annots']
                    for annot in annots:
                        annot_obj = annot.get_object()
                        if '/A' in annot_obj and '/URI' in annot_obj['/A']:
                            uri = annot_obj['/A']['/URI']
                            if uri:
                                links.add(uri)
    except Exception as e:
        log_error(f"Failed to extract links from PDF {file_path}: {e}")
    print(f"Extracted {len(links)} links from file {file_path}")
    return links

def extract_links_from_docx(file_path):
    links = set()
    try:
        doc = docx.Document(file_path)
        for paragraph in doc.paragraphs:
            for run in paragraph.runs:
                if run.hyperlink:
                    links.add(run.hyperlink.target)
    except Exception as e:
        log_error(f"Failed to extract links from DOCX {file_path}: {e}")
    print(f"Extracted {len(links)} links from file {file_path}")
    return links

def extract_links_from_csv(file_path):
    links = set()
    try:
        df = pd.read_csv(file_path)
        for col in df.columns:
            for val in df[col].values:
                if isinstance(val, str) and val.startswith('http'):
                    links.add(val)
    except Exception as e:
        log_error(f"Failed to extract links from CSV {file_path}: {e}")
    print(f"Extracted {len(links)} links from file {file_path}")
    return links

def extract_links_from_excel(file_path):
    links = set()
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        links.add(cell.hyperlink.target)
                    elif isinstance(cell.value, str) and cell.value.startswith('http'):
                        links.add(cell.value)
    except Exception as e:
        log_error(f"Failed to extract links from Excel {file_path}: {e}")
    print(f"Extracted {len(links)} links from file {file_path}")
    return links


def extract_links_from_file(file_path):
    try:
        if file_path.endswith('.pdf'):
            return extract_links_from_pdf(file_path)
        elif file_path.endswith('.docx'):
            return extract_links_from_docx(file_path)
        elif file_path.endswith('.csv'):
            return extract_links_from_csv(file_path)
        elif file_path.endswith('.xls') or file_path.endswith('.xlsx'):
            return extract_links_from_excel(file_path)
        else:
            return set()
    except Exception as e:
        log_error(f"Failed to extract text from {file_path}: {e}")
        return set()
    

def download_file(url, file_path):
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # Raise HTTPError for bad responses
        with open(file_path, 'wb') as file:
            file.write(response.content)
        print(f"Downloaded file from url : {url} and storing it at {file_path}")
        return True
    except Exception as e:
        failed_download_links_file = 'links/failed_download_links_file.txt'
        save_link(url, failed_download_links_file)
        log_error(f"Failed to download file {url}: {e}")
        return False
    

